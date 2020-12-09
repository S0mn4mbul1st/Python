# https://www.kaggle.com/new-york-city/nyc-property-sales

import collections
import csv
import sys
import openpyxl
from openpyxl.styles import Font

rows = []
column = []

def read_csv():
    path = sys.argv[1]
    try:
        with open(path) as dataFile:
            dictReader = csv.DictReader(dataFile)
            for row in dictReader:
                rows.append(row)
        return rows
    except FileNotFoundError:
        print("Invalid Path")
        pass


def get_avg_block_count():
    avg_block_amount = 0
    counter = 1
    for row in rows:
        counter += 1
        avg_block_amount += int(row['BLOCK'])
    return avg_block_amount / counter


def get_tot_lot_according_to_neighborhood(city):
    tot_lot = 0
    for row in rows:
        if city == row['NEIGHBORHOOD']:
            tot_lot += int(row['LOT'])
    return tot_lot


def get_unique_values_by_column(list):
    uniqueWords = []
    for i in list:
        if not i in uniqueWords:
            uniqueWords.append(i)
    return uniqueWords


def number_of_items():
    count = 0
    for i in rows:
        count += 1

    return count


stat = int(get_avg_block_count())
agg = get_tot_lot_according_to_neighborhood('CHELSEA')
summary = number_of_items()


def create_workbook():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'My sheet'
    sheet2 = workbook.create_sheet(title='Next sheet')
    workbook.save('mySheet.xlsx')


def open_workbook():
    workbook = openpyxl.load_workbook('mySheet.xlsx')
    print (workbook.sheetnames)
    sheet2 = workbook['Mysheet']

def get_outputs():
    return (int(get_avg_block_count()), get_tot_lot_according_to_neighborhood('CHELSEA'), number_of_items())

def save_to_excel():
    print("done")
    if sys.argv[2] == "-o":
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "MySheet"
        sheet.append((
            'BOROUGH',
            'NEIGHBORHOOD',
            'BUILDING CLASS CATEGORY',
            'TAX CLASS AT PRESENT',
            'BLOCK',
            'LOT',
            'EASE-MENT',
            'BUILDING CLASS AT PRESENT',
            'ADDRESS',
            'APARTMENT NUMBER',
            'ZIP CODE',
            'RESIDENTIAL UNITS',
            'COMMERCIAL UNITS',
            'TOTAL UNITS',
            'LAND SQUARE FEET',
            'GROSS SQUARE FEET',
            'YEAR BUILT',
            'TAX CLASS AT TIME OF SALE',
            'BUILDING CLASS AT TIME OF SALE',
            'SALE PRICE',
            'SALE DATE',
            ))

        sheet.append(get_outputs())

        print("done")
        sheet.cell(1, 1).font = Font(color='ff0000', size=15)
        sheet.cell(1, 2).font = Font(color='ff0000', size=15)
        sheet.cell(1, 3).font = Font(color='ff0000', size=15)
        sheet.cell(1, 4).font = Font(color='ff0000', size=15,
                bold=True)
        sheet.cell(1, 5).font = Font(color='ff0000', size=15,
                bold=False)

        workbook.save(sys.argv[3])

if __name__ == '__main__':
    read_csv()
    save_to_excel()